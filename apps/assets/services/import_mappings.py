from openpyxl import load_workbook
from datetime import datetime, date, time
from decimal import Decimal


def make_json_safe(value):
    """
    Convert Excel/python values into JSON-safe primitives
    before saving into JSONField.
    """
    if value is None:
        return None

    if isinstance(value, str):
        value = value.strip()
        return value or None

    if isinstance(value, (datetime, date, time)):
        return value.isoformat()

    if isinstance(value, Decimal):
        return float(value)

    if isinstance(value, (int, float, bool)):
        return value

    return str(value).strip()


def normalize_cell_value(value):
    return make_json_safe(value)


def build_normalized_rows(file_obj, mappings):
    """
    Reads the first sheet and applies source-column -> target-field mappings.
    Returns normalized rows ready to store in AssetImportRow.
    """
    workbook = load_workbook(file_obj, read_only=True, data_only=True)
    worksheet = workbook.worksheets[0]
    rows = worksheet.iter_rows(values_only=True)

    try:
        header_row = next(rows)
    except StopIteration:
        workbook.close()
        return []

    headers = [str(cell).strip() if cell is not None else "" for cell in header_row]

    normalized_rows = []

    for index, row in enumerate(rows, start=2):
        raw_data = {}
        normalized_data = {}

        for col_index, header in enumerate(headers):
            value = row[col_index] if col_index < len(row) else None
            safe_value = make_json_safe(value)

            raw_data[header] = safe_value

            target_field = mappings.get(header)
            if not target_field:
                continue

            normalized_data[target_field] = normalize_cell_value(value)

        normalized_rows.append(
            {
                "row_number": index,
                "raw_data": raw_data,
                "normalized_data": normalized_data,
            }
        )

    workbook.close()
    return normalized_rows