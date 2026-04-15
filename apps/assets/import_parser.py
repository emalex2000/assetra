from openpyxl import load_workbook
from datetime import datetime, date, time
from decimal import Decimal


def make_json_safe(value):
    if value is None:
        return ""

    if isinstance(value, str):
        return value.strip()

    if isinstance(value, (datetime, date, time)):
        return value.isoformat()

    if isinstance(value, Decimal):
        return float(value)

    if isinstance(value, (int, float, bool)):
        return value

    return str(value)


def extract_excel_metadata(file_obj):
    workbook = load_workbook(file_obj, read_only=True, data_only=True)
    worksheet = workbook.worksheets[0]
    rows = worksheet.iter_rows(values_only=True)

    try:
        header_rows = next(rows)
    except StopIteration:
        workbook.close()
        return {
            "sheet_name": worksheet.title,
            "headers": [],
            "total_rows": 0,
            "preview_rows": [],
        }

    headers = [str(cell).strip() if cell is not None else "" for cell in header_rows]
    preview_rows = []
    total_rows = 0

    for row in rows:
        total_rows += 1

        if len(preview_rows) < 5:
            preview_rows.append([make_json_safe(cell) for cell in row])

    workbook.close()

    return {
        "sheet_name": worksheet.title,
        "headers": headers,
        "total_rows": total_rows,
        "preview_rows": preview_rows,
    }